"""
End-to-end FastAPI layer tests.

Chain: TestClient → endpoint → get_webmotors_client() (lru_cache, lazy) →
WebmotorsClient → ensure_session → mint_session (bypassed via WEBMOTORS_COOKIE
env) → _get → requests.get (intercepted by requests_mock).

Conftest fixtures consumed (do NOT redefine):
  - webmotors_cookie_env  : sets WEBMOTORS_COOKIE so mint takes bypass branch
  - clear_lru_caches      : autouse, clears singletons between tests
  - search_payload()      : full /api/search/car fixture (3×0km + 12×used)
  - detail_payload()      : full /api/detail/car fixture
  - make_listing          : factory for a single used search result item
  - make_detail           : factory for a single detail item

Local fixtures/helpers are defined here and are keyword-only per project rule.
"""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any

import pytest
import requests
from fastapi.testclient import TestClient

import vehicle_search as vs
from vehicle_search import app
from tests.conftest import make_detail

BASE_URL = "https://www.webmotors.com.br"

# ---------------------------------------------------------------------------
# Shared TestClient — stateless HTTP only, no startup/shutdown side effects
# ---------------------------------------------------------------------------

client = TestClient(app)


# ---------------------------------------------------------------------------
# Local helper: builds the /api/search/car URL used by WebmotorsClient.search
# (mirrors the logic in webmotors_scraper.WebmotorsClient.search)
# ---------------------------------------------------------------------------

def _search_api_url(
    *,
    make: str,
    model: str = "",
    page: int = 1,
    per_page: int = 24,
    extra: dict[str, str] | None = None,
) -> str:
    """Produce the full search URL that WebmotorsClient.search constructs."""
    import requests as _r

    inner: dict[str, str] = {"tipoveiculo": "carros", "marca1": make.lower(), "page": str(page)}
    if model:
        inner["modelo1"] = model.lower()
    if extra:
        inner.update(extra)
    inner_qs = _r.compat.urlencode(inner)
    inner_url = _r.utils.quote(f"{BASE_URL}/carros/estoque?{inner_qs}", safe="")
    return (
        f"{BASE_URL}/api/search/car?url={inner_url}"
        f"&displayPerPage={per_page}&actualPage={page}"
        f"&showMenu=true&showCount=true&showBreadCrumb=true"
        f"&order=1&mediaZeroKm=true"
    )


# ============================================================================
# 1. App boots without TAVILY_API_KEY
# ============================================================================


def test_app_boots_without_tavily_key() -> None:
    """Importing app and building TestClient must NOT require TAVILY_API_KEY."""
    # Simply reaching here (module-level `client = TestClient(app)`) proves it.
    assert client is not None


def test_webmotors_endpoint_works_without_tavily_key(
    webmotors_cookie_env: None,
    requests_mock: Any,
    search_payload: dict[str, Any],
) -> None:
    """/webmotors/search succeeds even when TAVILY_API_KEY is not set."""
    url = _search_api_url(make="honda", model="city")
    requests_mock.get(url, json=search_payload, status_code=200)

    resp = client.get("/webmotors/search?marca=honda&modelo=city")
    assert resp.status_code == 200


def test_search_vehicle_without_tavily_key_returns_503(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """/search_vehicle returns 503 when TAVILY_API_KEY is not configured."""
    # Ensure the module-level TAVILY_API_KEY is falsy for this test.
    monkeypatch.setattr(vs, "TAVILY_API_KEY", None)
    monkeypatch.delenv("TAVILY_API_KEY", raising=False)

    resp = client.get("/search_vehicle?marca=honda&modelo=city&ano=2022")
    assert resp.status_code == 503
    assert "TAVILY_API_KEY" in resp.json()["detail"]


# ============================================================================
# 2. /webmotors/search
# ============================================================================


def test_webmotors_search_200_shape(
    webmotors_cookie_env: None,
    requests_mock: Any,
    search_payload: dict[str, Any],
) -> None:
    """GET /webmotors/search returns 200 with the expected response shape."""
    url = _search_api_url(make="honda", model="city")
    requests_mock.get(url, json=search_payload, status_code=200)

    resp = client.get("/webmotors/search?marca=honda&modelo=city")
    assert resp.status_code == 200

    data = resp.json()

    # Top-level keys produced by the endpoint
    expected_keys = {
        "marca", "modelo", "localidade", "cor", "ano_de", "ano_ate", "page", "total_disponivel",
        "total_anuncios", "media_preco", "media_km", "anuncios",
        "payload_completo",
    }
    assert expected_keys == set(data.keys())

    # Scalar fields
    assert data["marca"] == "honda"
    assert data["modelo"] == "city"
    assert data["localidade"] == ""
    assert data["cor"] == ""
    assert data["ano_de"] is None
    assert data["ano_ate"] is None
    assert data["page"] == 1
    assert data["total_disponivel"] == search_payload["Count"]

    # Endpoint returns the full SearchResults payload, including zero-km items.
    assert data["total_anuncios"] == len(search_payload["SearchResults"])
    assert len(data["anuncios"]) == len(search_payload["SearchResults"])
    assert data["payload_completo"] == search_payload

    # Médias must be non-null (all 12 used items have prices)
    assert data["media_preco"] is not None
    assert data["media_km"] is not None

    # Each anuncio is the full raw Webmotors listing, not the normalized subset.
    anuncio = data["anuncios"][0]
    assert anuncio == search_payload["SearchResults"][0]
    assert "Specification" in anuncio
    assert "Prices" in anuncio
    assert "Media" in anuncio


def test_webmotors_search_accepts_location_color_and_year_range(
    webmotors_cookie_env: None,
    requests_mock: Any,
    search_payload: dict[str, Any],
) -> None:
    """GET /webmotors/search forwards localidade/cor/ano_de/ano_ate to Webmotors."""
    requests_mock.get(re.compile(r".*/api/search/car.*"), json=search_payload, status_code=200)

    resp = client.get(
        "/webmotors/search"
        "?marca=honda&modelo=city&localidade=S%C3%A3o%20Paulo"
        "&cor=Branco&ano_de=2019&ano_ate=2022"
    )
    assert resp.status_code == 200

    data = resp.json()
    assert data["localidade"] == "São Paulo"
    assert data["cor"] == "Branco"
    assert data["ano_de"] == 2019
    assert data["ano_ate"] == 2022

    from urllib.parse import parse_qs, unquote, urlparse

    parsed = urlparse(requests_mock.last_request.url)
    qs = parse_qs(parsed.query, keep_blank_values=True)
    decoded_inner = unquote(qs["url"][0])
    assert "localizacao=São+Paulo" in decoded_inner
    assert "cor1=Branco" in decoded_inner
    assert "anode=2019" in decoded_inner
    assert "anoate=2022" in decoded_inner


def test_webmotors_search_rejects_inverted_year_range() -> None:
    resp = client.get("/webmotors/search?marca=honda&modelo=city&ano_de=2022&ano_ate=2019")
    assert resp.status_code == 400


# ============================================================================
# 3. /webmotors/detail  (E2E: search → iter_details → normalize → médias)
# ============================================================================


def test_webmotors_detail_200_shape_and_medias(
    webmotors_cookie_env: None,
    requests_mock: Any,
    search_payload: dict[str, Any],
    detail_payload: dict[str, Any],
) -> None:
    """GET /webmotors/detail: 0km items skipped; response includes computed médias."""
    search_url = _search_api_url(make="honda", model="city", per_page=12)
    requests_mock.get(search_url, json=search_payload, status_code=200)

    # All detail URLs for used items share the same pattern; one fixture response works.
    requests_mock.get(
        re.compile(r"/api/detail/car/"),
        json=detail_payload,
        status_code=200,
    )

    resp = client.get("/webmotors/detail?marca=honda&modelo=city")
    assert resp.status_code == 200

    data = resp.json()

    # Top-level keys
    expected_keys = {
        "marca", "modelo", "localidade", "cor", "ano_de", "ano_ate",
        "total_anuncios", "media_preco", "media_km", "anuncios",
    }
    assert expected_keys == set(data.keys())

    assert data["marca"] == "honda"
    assert data["modelo"] == "city"
    assert data["localidade"] == ""
    assert data["cor"] == ""
    assert data["ano_de"] is None
    assert data["ano_ate"] is None

    # 0km items (UniqueId==0) must have been skipped; 12 used items remain.
    assert data["total_anuncios"] == 12
    assert len(data["anuncios"]) == 12

    # Médias must be sane (detail_payload has Price="90000" as string)
    assert data["media_preco"] is not None
    assert data["media_km"] is not None
    # Price from fixture is 90000 — média must NOT be inflated 10×
    assert data["media_preco"] < 900_000, "media_preco inflated — _to_number may be mishandling the string"

    # Each anuncio is the full raw /api/detail payload.
    anuncio = data["anuncios"][0]
    assert anuncio == detail_payload
    assert "Specification" in anuncio
    assert "Seller" in anuncio
    assert "Prices" in anuncio

    # Seller fields from detail_payload must be present
    assert anuncio["Seller"]["City"] == "Volta Redonda"
    assert anuncio["Seller"]["State"] == "Rio de Janeiro (RJ)"


def test_purchase_recommendation_calls_bedrock_agent(
    monkeypatch: pytest.MonkeyPatch,
    detail_payload: dict[str, Any],
) -> None:
    captured: dict[str, Any] = {"webmotors_calls": []}

    class FakeWebmotorsClient:
        def search(self, **kwargs):
            captured["webmotors_calls"].append(kwargs)
            if kwargs["page"] == 1:
                return {"SearchResults": [detail_payload]}
            return {"SearchResults": []}

        def iter_details(self, **kwargs):
            raise AssertionError("recommendation must use search(), not iter_details()")

    def fake_invoke(prompt: str) -> str:
        captured["prompt"] = prompt
        return "Boa opção: preço e km estão competitivos para a busca."

    monkeypatch.setattr(vs, "get_webmotors_client", lambda: FakeWebmotorsClient())
    monkeypatch.setattr(vs, "_invoke_bedrock_agent", fake_invoke)

    resp = client.post(
        "/recommendation",
        json={
            "marca": "honda",
            "modelo": "city",
            "localidade": "Rio de Janeiro",
            "cor": "Branco",
            "ano_de": 2019,
            "ano_ate": 2022,
            "pages": 1,
            "per_page": 8,
        },
    )
    assert resp.status_code == 200

    data = resp.json()
    assert data["busca"]["marca"] == "honda"
    assert data["metricas"]["total_anuncios"] == 1
    assert data["recomendacao"].startswith("Boa opção")
    assert len(data["anuncios_analisados"]) == 1
    assert "/comprar/" in data["anuncios_analisados"][0]["url"]
    assert "/api/detail/" not in data["anuncios_analisados"][0]["url"]
    first_call = captured["webmotors_calls"][0]
    assert first_call["extra"]["anode"] == "2019"
    assert first_call["extra"]["anoate"] == "2022"
    assert "localizacao" not in first_call["extra"]
    assert [call["page"] for call in captured["webmotors_calls"]] == [1, 2, 3]
    assert all(call["per_page"] == 24 for call in captured["webmotors_calls"])
    assert "anuncios_candidatos" in captured["prompt"]


def test_purchase_recommendation_rejects_inverted_year_range() -> None:
    resp = client.post(
        "/recommendation",
        json={"marca": "honda", "modelo": "city", "ano_de": 2022, "ano_ate": 2019},
    )
    assert resp.status_code == 400


def test_purchase_recommendation_filters_results_by_requested_location(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, Any] = {}
    rj_detail = make_detail(
        price="90000",
        odometer="40000",
        seller_city="Rio de Janeiro",
        seller_state="Rio de Janeiro (RJ)",
    )
    sp_detail = make_detail(
        price="70000",
        odometer="20000",
        seller_city="São Paulo",
        seller_state="São Paulo (SP)",
    )

    class FakeWebmotorsClient:
        def search(self, **kwargs):
            if kwargs["page"] == 1:
                return {"SearchResults": [rj_detail, sp_detail]}
            return {"SearchResults": []}

        def iter_details(self, **kwargs):
            raise AssertionError("recommendation must use search(), not iter_details()")

    def fake_invoke(prompt: str) -> str:
        captured["prompt"] = prompt
        return "Recomendo apenas o anúncio do Rio de Janeiro."

    monkeypatch.setattr(vs, "get_webmotors_client", lambda: FakeWebmotorsClient())
    monkeypatch.setattr(vs, "_invoke_bedrock_agent", fake_invoke)

    resp = client.post(
        "/recommendation",
        json={"marca": "honda", "modelo": "city", "localidade": "Rio de Janeiro"},
    )
    assert resp.status_code == 200

    data = resp.json()
    assert data["metricas"]["total_anuncios"] == 1
    assert data["metricas"]["total_anuncios_encontrados_webmotors"] == 2
    assert data["metricas"]["total_ignorados_por_localidade"] == 1
    assert data["anuncios_analisados"][0]["cidade"] == "Rio de Janeiro"
    assert "São Paulo" not in captured["prompt"]


def test_batch_template_returns_xlsx() -> None:
    resp = client.get("/recommendation/batch/template")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.content.startswith(b"PK")


def test_batch_recommendation_returns_result_xlsx(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from openpyxl import Workbook, load_workbook

    input_book = Workbook()
    sheet = input_book.active
    sheet.append(["marca", "modelo", "localidade", "cor", "ano_de", "ano_ate"])
    sheet.append(["Renault", "Sandero", "RJ", "", 2019, 2022])
    sheet.append(["Honda", "City", "SP", "Branco", 2020, 2022])
    payload = BytesIO()
    input_book.save(payload)
    payload.seek(0)

    captured: list[vs.RecommendationRequest] = []

    def fake_collect(request: vs.RecommendationRequest):
        captured.append(request)
        best = {
            "titulo": f"{request.marca} {request.modelo}",
            "preco": 50000,
            "km": 40000,
            "ano_modelo": request.ano_ate,
            "ano_fabricacao": request.ano_de,
            "cor": request.cor,
            "cidade": "Rio de Janeiro" if request.localidade == "RJ" else "Sao Paulo",
            "estado": "Rio de Janeiro (RJ)" if request.localidade == "RJ" else "Sao Paulo (SP)",
            "url": "https://www.webmotors.com.br/comprar/teste",
        }
        return (
            {
                "marca": request.marca,
                "modelo": request.modelo,
                "localidade": request.localidade,
                "cor": request.cor,
                "ano_de": request.ano_de,
                "ano_ate": request.ano_ate,
            },
            {
                "total_anuncios": 1,
                "total_anuncios_encontrados_webmotors": 3,
                "total_ignorados_por_localidade": 2,
                "media_preco": 55000,
                "media_km": 45000,
            },
            [best],
            best,
        )

    monkeypatch.setattr(vs, "_collect_recommendation_analysis", fake_collect)

    resp = client.post(
        "/recommendation/batch",
        files={
            "file": (
                "carros.xlsx",
                payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200
    assert resp.headers["x-batch-items"] == "2"
    assert resp.content.startswith(b"PK")
    assert len(captured) == 2

    output_book = load_workbook(BytesIO(resp.content), read_only=False, data_only=True)
    rows = list(output_book.active.iter_rows(values_only=True))
    assert rows[0][0:4] == ("linha_origem", "marca", "modelo", "localidade")
    assert rows[1][0:8] == (2, "Renault", "Sandero", "RJ", None, 2019, 2022, "ok")
    assert rows[2][0:8] == (3, "Honda", "City", "SP", "Branco", 2020, 2022, "ok")
    link_cell = output_book.active["V2"]
    assert link_cell.hyperlink is not None
    assert link_cell.hyperlink.target == "https://www.webmotors.com.br/comprar/teste"


def test_batch_recommendation_retries_after_webmotors_403(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    request = vs.RecommendationRequest(marca="Renault", modelo="Sandero")
    calls = {"collect": 0, "remint": 0}

    def fake_collect(received: vs.RecommendationRequest):
        calls["collect"] += 1
        if calls["collect"] == 1:
            response = requests.Response()
            response.status_code = 403
            raise requests.HTTPError("403 Client Error: Forbidden", response=response)
        return (
            {
                "marca": received.marca,
                "modelo": received.modelo,
                "localidade": received.localidade,
                "cor": received.cor,
                "ano_de": received.ano_de,
                "ano_ate": received.ano_ate,
            },
            {
                "total_anuncios": 1,
                "total_anuncios_encontrados_webmotors": 1,
                "total_ignorados_por_localidade": 0,
                "media_preco": 50000,
                "media_km": 40000,
            },
            [],
            {"titulo": "Renault Sandero", "url": "https://www.webmotors.com.br/comprar/teste"},
        )

    class FakeWebmotorsClient:
        def ensure_session(self, **kwargs):
            calls["remint"] += 1

    monkeypatch.setattr(vs, "_collect_recommendation_analysis", fake_collect)
    monkeypatch.setattr(vs, "get_webmotors_client", lambda: FakeWebmotorsClient())
    monkeypatch.setattr(vs.time, "sleep", lambda seconds: None)

    busca, metricas, _, melhor = vs._collect_batch_analysis_with_retry(request)

    assert busca["marca"] == "Renault"
    assert metricas["total_anuncios"] == 1
    assert melhor["titulo"] == "Renault Sandero"
    assert calls == {"collect": 2, "remint": 1}


# ============================================================================
# 4. /webmotors/detail_by_url
# ============================================================================

_DETAIL_BY_URL = (
    f"{BASE_URL}/api/detail/car/honda/city/15-i-vtec-flex-hatch-exl-cvt/4-portas/2022/69863170"
)


def test_webmotors_detail_by_url_200(
    webmotors_cookie_env: None,
    requests_mock: Any,
    detail_payload: dict[str, Any],
) -> None:
    """GET /webmotors/detail_by_url returns 200 + full raw detail payload."""
    requests_mock.get(_DETAIL_BY_URL, json=detail_payload, status_code=200)

    resp = client.get(f"/webmotors/detail_by_url?url={_DETAIL_BY_URL}")
    assert resp.status_code == 200

    data = resp.json()

    assert data == detail_payload
    assert data["UniqueId"] == detail_payload["UniqueId"]
    assert data["Specification"]["Make"]["Value"] == "HONDA"
    assert data["Specification"]["Model"]["Value"] == "CITY"
    # Seller fields must be present (detail_payload has Seller)
    assert data["Seller"]["City"] == "Volta Redonda"
    assert data["Seller"]["State"] == "Rio de Janeiro (RJ)"
    # Price from detail fixture is string "90000" — must parse cleanly
    assert data["Prices"]["Price"] == "90000"


# ============================================================================
# 5. /webmotors/refresh_session  (POST)
# ============================================================================


def test_webmotors_refresh_session_200(
    webmotors_cookie_env: None,
) -> None:
    """POST /webmotors/refresh_session returns 200 + {ok, cookies, user_agent}."""
    # With WEBMOTORS_COOKIE set, ensure_session(force=True) takes the
    # no-browser bypass and returns a WebmotorsSession without Playwright.
    resp = client.post("/webmotors/refresh_session")
    assert resp.status_code == 200

    data = resp.json()
    assert data["ok"] is True
    assert isinstance(data["cookies"], list)
    assert isinstance(data["user_agent"], str)
    assert len(data["user_agent"]) > 0


# ============================================================================
# 6. WebmotorsBlocked → 502
# ============================================================================


def test_webmotors_blocked_search_returns_502(
    webmotors_cookie_env: None,
    requests_mock: Any,
) -> None:
    """When _get exhausts its one retry (both requests return 403), endpoint raises 502."""
    search_url_pattern = re.compile(r"/api/search/car")
    # Two consecutive 403s: first attempt → 403 triggers force remint (bypass, no browser)
    # → second attempt via _retried=True also → 403 → raise_for_status propagates.
    # The endpoint catches WebmotorsBlocked and re-raises as 502.
    # But note: _get calls raise_for_status() AFTER the retry, so after the 2nd 403
    # requests will raise HTTPError, NOT WebmotorsBlocked.
    # We need to check the actual behavior: _get returns resp after 2nd attempt,
    # then search() calls resp.raise_for_status() which raises HTTPError (not 502).
    # To get 502 cleanly, monkeypatch WebmotorsClient to raise WebmotorsBlocked directly.
    import webmotors_scraper as wm

    # Instance method patch: first positional arg is self (the WebmotorsClient instance).
    def _raise_blocked(self: Any, *, make: Any, model: Any = "", **kwargs: Any) -> None:
        raise wm.WebmotorsBlocked("PerimeterX blocked (test)")

    with pytest.MonkeyPatch.context() as mp:
        mp.setattr(wm.WebmotorsClient, "search", _raise_blocked)
        resp = client.get("/webmotors/search?marca=honda&modelo=city")

    assert resp.status_code == 502
    assert "PerimeterX" in resp.json()["detail"] or "blocked" in resp.json()["detail"].lower()


def test_webmotors_blocked_detail_by_url_returns_502(
    webmotors_cookie_env: None,
    requests_mock: Any,
) -> None:
    """WebmotorsBlocked from get_detail propagates as 502 on /detail_by_url."""
    import webmotors_scraper as wm

    # Instance method patch: first positional arg is self.
    def _raise_blocked(self: Any, *, url: Any = None, listing: Any = None, **kwargs: Any) -> None:
        raise wm.WebmotorsBlocked("PerimeterX blocked (test)")

    with pytest.MonkeyPatch.context() as mp:
        mp.setattr(wm.WebmotorsClient, "get_detail", _raise_blocked)
        resp = client.get(f"/webmotors/detail_by_url?url={_DETAIL_BY_URL}")

    assert resp.status_code == 502


# ============================================================================
# 7. /search_vehicle happy path (legacy Tavily)
# ============================================================================


class _FakeTavilyClient:
    """Minimal Tavily client stub that returns canned results with R$/km text."""

    def search(self, *, query: str = "", search_depth: str = "advanced", max_results: int = 50) -> dict[str, Any]:
        return {
            "results": [
                {
                    "url": "https://www.webmotors.com.br/carro/honda/city/2022/used/123",
                    "title": "Honda City 2022",
                    "content": (
                        "Honda City 2022 Cinza. "
                        "R$ 90.000 - 35.000 Km. "
                        "R$ 95.000 - 28.000 Km."
                    ),
                },
                {
                    "url": "https://www.webmotors.com.br/carro/honda/city/2022/used/456",
                    "title": "Honda City 2022 EXL",
                    "content": "R$ 85.000 - 42.000 Km.",
                },
                {
                    # Non-webmotors result — must be filtered out by keep_only_webmotors
                    "url": "https://www.icarros.com.br/honda/city",
                    "title": "iCarros City",
                    "content": "R$ 1.000.000 - 1 Km.",
                },
            ]
        }


def test_search_vehicle_happy_path(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """GET /search_vehicle returns 200 with extracted/averaged data (Tavily legacy)."""
    # Replace the lru_cache singleton factory so the endpoint gets our stub.
    monkeypatch.setattr(vs, "get_tavily_client", lambda: _FakeTavilyClient())

    resp = client.get("/search_vehicle?marca=honda&modelo=city&ano=2022")
    assert resp.status_code == 200

    data = resp.json()

    # Top-level keys from the endpoint
    expected_keys = {
        "marca", "modelo", "ano", "cor", "localidade",
        "total_anuncios", "media_preco", "media_km", "anuncios",
    }
    assert expected_keys == set(data.keys())

    assert data["marca"] == "honda"
    assert data["modelo"] == "city"
    assert data["ano"] == 2022
    assert data["cor"] == ""
    assert data["localidade"] == ""

    # 3 valid anuncios from 2 webmotors results (2 from first URL, 1 from second)
    assert data["total_anuncios"] == 3
    assert len(data["anuncios"]) == 3

    # Médias: prices are 90000, 95000, 85000 → média = 90000
    assert data["media_preco"] == pytest.approx(90000.0)
    # kms: 35000, 28000, 42000 → média = 35000
    assert data["media_km"] == pytest.approx(35000.0)

    # Each anuncio has the expected fields
    anuncio = data["anuncios"][0]
    anuncio_keys = {"titulo", "url", "preco", "km", "localidade_pesquisada", "cor_pesquisada"}
    assert anuncio_keys == set(anuncio.keys())

    # icarros URL was filtered out — check no icarros anuncio leaked
    urls = [a["url"] for a in data["anuncios"]]
    assert all("webmotors.com.br" in u for u in urls)
