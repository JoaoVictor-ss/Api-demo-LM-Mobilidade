from fastapi import FastAPI, HTTPException, Query, File, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from tavily import TavilyClient
from dotenv import load_dotenv
from pathlib import Path
from functools import lru_cache
from pydantic import BaseModel, Field
from typing import Any
from io import BytesIO
import json
import os
import re
import time
import unicodedata
import uuid

import webmotors_scraper as wm

# ==========================================
# Carrega .env
# ==========================================

BASE_DIR = Path(__file__).resolve().parent

load_dotenv(BASE_DIR / ".env")

TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")
BEDROCK_AGENT_ID = os.getenv("BEDROCK_AGENT_ID")
BEDROCK_AGENT_ALIAS_ID = os.getenv("BEDROCK_AGENT_ALIAS_ID")
AWS_REGION = os.getenv("AWS_REGION") or os.getenv("AWS_DEFAULT_REGION") or "us-east-1"

# ==========================================
# FastAPI
# ==========================================

app = FastAPI(title="LM Mobilidade - Webmotors")

STATIC_DIR = BASE_DIR / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


class RecommendationRequest(BaseModel):
    marca: str = Field(..., min_length=1)
    modelo: str = ""
    localidade: str = ""
    cor: str = ""
    ano_de: int | None = None
    ano_ate: int | None = None
    pages: int = Field(1, ge=1, le=3)
    per_page: int = Field(8, ge=1, le=24)


class RecommendationResponse(BaseModel):
    busca: dict[str, Any]
    metricas: dict[str, Any]
    melhor_opcao_local: dict[str, Any] | None
    recomendacao: str
    anuncios_analisados: list[dict[str, Any]]


BATCH_MAX_ROWS = 20
BATCH_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# Tavily é lazy: o app sobe mesmo sem a key (os endpoints /webmotors não
# dependem dela). A key só é exigida quando /search_vehicle é chamado.
@lru_cache
def get_tavily_client():
    if not TAVILY_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="TAVILY_API_KEY não configurada no .env",
        )
    return TavilyClient(api_key=TAVILY_API_KEY)


def get_bedrock_agent_client():
    region, _, _ = get_bedrock_config()
    try:
        import boto3
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="Dependência boto3 não instalada. Rode: pip install boto3",
        ) from exc
    return boto3.client("bedrock-agent-runtime", region_name=region)


def get_bedrock_config() -> tuple[str, str, str]:
    load_dotenv(BASE_DIR / ".env", override=True)
    region = os.getenv("AWS_REGION") or os.getenv("AWS_DEFAULT_REGION") or "us-east-1"
    agent_id = os.getenv("BEDROCK_AGENT_ID")
    agent_alias_id = os.getenv("BEDROCK_AGENT_ALIAS_ID")
    if not agent_id or not agent_alias_id:
        raise HTTPException(
            status_code=503,
            detail="BEDROCK_AGENT_ID e BEDROCK_AGENT_ALIAS_ID precisam estar configurados no .env",
        )
    return region, agent_id, agent_alias_id


# Cliente Webmotors compartilhado entre requests (reusa a sessão PerimeterX
# em cache; só re-minta quando expira ou toma 403).
@lru_cache
def get_webmotors_client() -> wm.WebmotorsClient:
    return wm.WebmotorsClient()


@app.get("/", include_in_schema=False)
def ui_home():
    index_path = STATIC_DIR / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=404, detail="UI não encontrada")
    return FileResponse(index_path)


# ==========================================
# Mantém apenas Webmotors
# ==========================================

def keep_only_webmotors(results):

    return [
        r
        for r in results
        if "webmotors.com.br" in r.get("url", "")
    ]

# ==========================================
# Extrai preço e km
# ==========================================

def extract_vehicle_data(text):

    preco_pattern = r"R\$\s?([\d\.]+)"
    km_pattern = r"([\d\.]+)\s?Km"

    precos = re.findall(
        preco_pattern,
        text,
        flags=re.IGNORECASE
    )

    kms = re.findall(
        km_pattern,
        text,
        flags=re.IGNORECASE
    )

    anuncios = []

    total = min(
        len(precos),
        len(kms)
    )

    for i in range(total):

        try:

            preco = int(
                precos[i].replace(".", "")
            )

            km = int(
                kms[i].replace(".", "")
            )

            anuncios.append({
                "preco": preco,
                "km": km
            })

        except:
            pass

    return anuncios


# ==========================================
# Médias (preço/km) sobre anúncios normalizados
# ==========================================

def _to_number(value):
    # A Webmotors devolve número nativo na busca (115900.0) e string no detail
    # ("115900"). Tavily já entrega int. Tratamos os três sem inflar o valor.
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[^\d.,]", "", str(value).strip())
    if not s:
        return None
    if "," in s:  # pt-BR: ponto é milhar, vírgula é decimal
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def calcular_medias(anuncios, *, campo_preco="preco", campo_km="km"):
    precos = [n for n in (_to_number(a.get(campo_preco)) for a in anuncios) if n]
    kms = [n for n in (_to_number(a.get(campo_km)) for a in anuncios) if n]
    return {
        "media_preco": round(sum(precos) / len(precos), 2) if precos else None,
        "media_km": round(sum(kms) / len(kms), 2) if kms else None,
    }


def _validate_year_range(ano_de: int | None, ano_ate: int | None) -> None:
    if ano_de is not None and ano_ate is not None and ano_de > ano_ate:
        raise HTTPException(status_code=400, detail="ano_de não pode ser maior que ano_ate")


def _build_webmotors_extra(
    *,
    localidade: str = "",
    cor: str = "",
    ano_de: int | None = None,
    ano_ate: int | None = None,
    include_localidade: bool = True,
) -> dict[str, str]:
    _validate_year_range(ano_de, ano_ate)
    return wm._build_search_extra(
        localidade=localidade if include_localidade else "",
        cor=cor,
        ano_de=ano_de,
        ano_ate=ano_ate,
    )


def _vehicle_url(raw: dict[str, Any]) -> str | None:
    public_url = _public_vehicle_url(raw)
    if public_url:
        return public_url

    direct_url = (
        raw.get("AdvertisementLink")
        or raw.get("Url")
        or raw.get("url")
        or raw.get("SEO", {}).get("Canonical")
    )
    if direct_url:
        return direct_url

    try:
        spec = raw["Specification"]
        return wm.build_detail_url(
            make=(spec.get("Make") or {})["Value"],
            model=(spec.get("Model") or {})["Value"],
            version=(spec.get("Version") or {})["Value"],
            doors=spec.get("NumberPorts", 4) or 4,
            year=spec.get("YearModel") or spec.get("YearFabrication"),
            unique_id=raw["UniqueId"],
        )
    except (KeyError, TypeError, ValueError):
        return None


def _year_slug(year_fabrication: Any, year_model: Any) -> str:
    fabrication = int(float(year_fabrication))
    model = int(float(year_model))
    return f"{fabrication}-{model}"


def _public_vehicle_url(raw: dict[str, Any]) -> str | None:
    try:
        spec = raw["Specification"]
        return (
            f"{wm.BASE_URL}/comprar/"
            f"{wm.slugify((spec.get('Make') or {})['Value'])}/"
            f"{wm.slugify((spec.get('Model') or {})['Value'])}/"
            f"{wm.slugify((spec.get('Version') or {})['Value'])}/"
            f"{int(float(spec.get('NumberPorts', 4) or 4))}-portas/"
            f"{_year_slug(spec.get('YearFabrication'), spec.get('YearModel'))}/"
            f"{raw['UniqueId']}"
        )
    except (KeyError, TypeError, ValueError):
        return None


def _normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


_STATE_ALIASES = {
    "ac": ("acre",),
    "al": ("alagoas",),
    "ap": ("amapa",),
    "am": ("amazonas",),
    "ba": ("bahia",),
    "ce": ("ceara",),
    "df": ("distrito federal", "brasilia"),
    "es": ("espirito santo",),
    "go": ("goias",),
    "ma": ("maranhao",),
    "mt": ("mato grosso",),
    "ms": ("mato grosso do sul",),
    "mg": ("minas gerais",),
    "pa": ("para",),
    "pb": ("paraiba",),
    "pr": ("parana",),
    "pe": ("pernambuco",),
    "pi": ("piaui",),
    "rj": ("rio de janeiro",),
    "rn": ("rio grande do norte",),
    "rs": ("rio grande do sul",),
    "ro": ("rondonia",),
    "rr": ("roraima",),
    "sc": ("santa catarina",),
    "sp": ("sao paulo",),
    "se": ("sergipe",),
    "to": ("tocantins",),
}


def _location_terms(localidade: str) -> list[str]:
    normalized = _normalize_text(localidade)
    if not normalized:
        return []

    terms = {normalized}
    parts = [part.strip() for part in re.split(r"[/,-]", localidade) if part.strip()]
    terms.update(_normalize_text(part) for part in parts)

    for uf, names in _STATE_ALIASES.items():
        if normalized == uf:
            terms.add(uf)
            terms.update(names)
        if normalized in names:
            terms.add(uf)

    return sorted(term for term in terms if term)


def _requested_state_codes(localidade: str) -> set[str]:
    normalized = _normalize_text(localidade)
    if not normalized:
        return set()

    parts = [normalized]
    parts.extend(_normalize_text(part) for part in re.split(r"[/,-]", localidade) if part.strip())
    tokens = normalized.split()
    parts.extend(tokens)

    codes: set[str] = set()
    for part in parts:
        if part in _STATE_ALIASES:
            codes.add(part)
        for uf, names in _STATE_ALIASES.items():
            if part in names:
                codes.add(uf)
    return codes


def _state_code_from_text(value: Any) -> str:
    normalized = _normalize_text(value)
    tokens = normalized.split()
    for token in reversed(tokens):
        if token in _STATE_ALIASES:
            return token
    return ""


def _matches_location(vehicle: dict[str, Any], localidade: str) -> bool:
    terms = _location_terms(localidade)
    if not terms:
        return True

    cidade = _normalize_text(vehicle.get("cidade"))
    estado = _normalize_text(vehicle.get("estado"))
    haystack = f"{cidade} {estado}"
    vehicle_uf = _state_code_from_text(vehicle.get("estado"))
    requested_ufs = _requested_state_codes(localidade)

    if requested_ufs:
        return vehicle_uf in requested_ufs

    for term in terms:
        if len(term) == 2 and term in _STATE_ALIASES:
            if term == vehicle_uf:
                return True
            continue
        if term == cidade or term == estado or term in haystack:
            return True
    return False


def _filter_by_location(
    vehicles: list[dict[str, Any]], localidade: str
) -> tuple[list[dict[str, Any]], int]:
    if not localidade:
        return vehicles, 0
    filtered = [vehicle for vehicle in vehicles if _matches_location(vehicle, localidade)]
    return filtered, len(vehicles) - len(filtered)


def _summarize_vehicle(raw: dict[str, Any]) -> dict[str, Any]:
    normalized = wm.normalize_detail(raw)
    return {
        "titulo": normalized.get("titulo"),
        "marca": normalized.get("marca"),
        "modelo": normalized.get("modelo"),
        "versao": normalized.get("versao"),
        "ano_modelo": normalized.get("ano_modelo"),
        "ano_fabricacao": normalized.get("ano_fabricacao"),
        "preco": _to_number(normalized.get("preco")),
        "km": _to_number(normalized.get("km")),
        "cor": normalized.get("cor"),
        "cambio": normalized.get("cambio"),
        "cidade": normalized.get("cidade"),
        "estado": normalized.get("estado"),
        "url": _vehicle_url(raw),
    }


def _rank_vehicle(vehicle: dict[str, Any], medias: dict[str, Any]) -> float:
    preco = vehicle.get("preco")
    km = vehicle.get("km")
    ano = _to_number(vehicle.get("ano_modelo")) or _to_number(vehicle.get("ano_fabricacao"))
    media_preco = medias.get("media_preco") or preco or 1
    media_km = medias.get("media_km") or km or 1
    preco_score = (preco / media_preco) if preco else 2
    km_score = (km / media_km) if km else 1
    ano_max = medias.get("ano_max") or ano or 1
    ano_score = ((ano_max - ano) / 10) if ano else 0.5
    return (preco_score * 0.55) + (km_score * 0.3) + (ano_score * 0.15)


def _prepare_analysis_payload(
    *,
    request: RecommendationRequest,
    details: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]], dict[str, Any] | None]:
    all_vehicles = [_summarize_vehicle(item) for item in details]
    vehicles, ignored_by_location = _filter_by_location(all_vehicles, request.localidade)
    medias = calcular_medias(vehicles)
    anos = [
        n
        for n in (
            _to_number(vehicle.get("ano_modelo")) or _to_number(vehicle.get("ano_fabricacao"))
            for vehicle in vehicles
        )
        if n
    ]
    if anos:
        medias["ano_max"] = max(anos)
    ranked = sorted(vehicles, key=lambda item: _rank_vehicle(item, medias))
    metricas = {
        "total_anuncios": len(vehicles),
        "total_anuncios_encontrados_webmotors": len(all_vehicles),
        "total_ignorados_por_localidade": ignored_by_location,
        "media_preco": medias["media_preco"],
        "media_km": medias["media_km"],
    }
    busca = {
        "marca": request.marca,
        "modelo": request.modelo,
        "localidade": request.localidade,
        "cor": request.cor,
        "ano_de": request.ano_de,
        "ano_ate": request.ano_ate,
    }
    return busca, metricas, ranked[:8], (ranked[0] if ranked else None)


def _build_agent_prompt(
    *,
    busca: dict[str, Any],
    metricas: dict[str, Any],
    anuncios: list[dict[str, Any]],
    melhor_opcao_local: dict[str, Any] | None,
) -> str:
    payload = {
        "objetivo": "Gerar recomendação de compra para o usuário com base em anúncios reais da Webmotors.",
        "instrucoes": [
            "Responda em portugues do Brasil.",
            "Use somente os anuncios_candidatos enviados no JSON.",
            "Os anuncios_candidatos ja foram filtrados pela localidade solicitada; nao use carros de outros estados ou cidades.",
            "Compare preco, quilometragem, ano, cor, localidade, versao e link quando disponiveis.",
            "Explique se a melhor opcao parece boa compra em relacao as medias da busca filtrada.",
            "Inclua o link do anuncio recomendado quando existir.",
            "Se nao houver anuncios na localidade solicitada, diga isso claramente e nao recomende carros de fora.",
            "Se os dados forem insuficientes, diga isso claramente e recomende proximos passos.",
            "Nao invente informacoes que nao estejam no JSON.",
        ],
        "busca": busca,
        "metricas": metricas,
        "melhor_opcao_calculada_pelo_backend": melhor_opcao_local,
        "anuncios_candidatos": anuncios,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _search_webmotors_pages(
    *,
    make: str,
    model: str,
    pages: int,
    per_page: int,
    extra: dict[str, str],
) -> list[dict[str, Any]]:
    client = get_webmotors_client()
    results: list[dict[str, Any]] = []
    for page in range(1, pages + 1):
        data = client.search(
            make=make,
            model=model,
            page=page,
            per_page=per_page,
            extra=extra,
        )
        results.extend(data.get("SearchResults", []))
    return results


def _collect_recommendation_analysis(
    request: RecommendationRequest,
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]], dict[str, Any] | None]:
    extra = _build_webmotors_extra(
        localidade=request.localidade,
        cor=request.cor,
        ano_de=request.ano_de,
        ano_ate=request.ano_ate,
        include_localidade=False,
    )
    search_pages = max(request.pages, 3) if request.localidade else request.pages
    search_per_page = max(request.per_page, 24) if request.localidade else request.per_page
    listings = _search_webmotors_pages(
        make=request.marca,
        model=request.modelo,
        pages=search_pages,
        per_page=search_per_page,
        extra=extra,
    )
    return _prepare_analysis_payload(request=request, details=listings)


def _invoke_bedrock_agent(prompt: str) -> str:
    client = get_bedrock_agent_client()
    _, agent_id, agent_alias_id = get_bedrock_config()
    response = client.invoke_agent(
        agentId=agent_id,
        agentAliasId=agent_alias_id,
        sessionId=str(uuid.uuid4()),
        inputText=prompt,
        enableTrace=False,
    )

    chunks: list[str] = []
    for event in response.get("completion", []):
        chunk = event.get("chunk")
        if not chunk:
            continue
        chunk_bytes = chunk.get("bytes")
        if isinstance(chunk_bytes, bytes):
            chunks.append(chunk_bytes.decode("utf-8"))
        elif chunk_bytes:
            chunks.append(str(chunk_bytes))

    answer = "".join(chunks).strip()
    if not answer:
        raise HTTPException(status_code=502, detail="Bedrock Agent não retornou texto")
    return answer


def _xlsx_tools():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="Dependencia openpyxl nao instalada. Rode: pip install openpyxl",
        ) from exc
    return Workbook, load_workbook, Font


def _canonical_excel_header(value: Any) -> str:
    normalized = _normalize_text(value).replace(" ", "_")
    aliases = {
        "marca": "marca",
        "make": "marca",
        "modelo": "modelo",
        "model": "modelo",
        "localidade": "localidade",
        "localizacao": "localidade",
        "cidade": "localidade",
        "estado": "localidade",
        "uf": "localidade",
        "cor": "cor",
        "ano": "ano",
        "ano_de": "ano_de",
        "ano_inicial": "ano_de",
        "de": "ano_de",
        "ano_ate": "ano_ate",
        "ano_final": "ano_ate",
        "ate": "ano_ate",
        "paginas": "pages",
        "pages": "pages",
        "itens_por_pagina": "per_page",
        "items_por_pagina": "per_page",
        "per_page": "per_page",
    }
    return aliases.get(normalized, "")


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _cell_int(value: Any) -> int | None:
    number = _to_number(value)
    return int(number) if number is not None else None


def _read_batch_requests(content: bytes) -> list[tuple[int, RecommendationRequest]]:
    _, load_workbook, _ = _xlsx_tools()
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo Excel vazio")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Nao foi possivel ler o Excel enviado") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise HTTPException(status_code=400, detail="Excel sem cabecalho") from exc

    headers: dict[str, int] = {}
    for index, value in enumerate(header_row):
        header = _canonical_excel_header(value)
        if header and header not in headers:
            headers[header] = index

    if "marca" not in headers:
        raise HTTPException(status_code=400, detail="A planilha precisa ter a coluna marca")

    batch: list[tuple[int, RecommendationRequest]] = []
    for excel_row_number, row in enumerate(rows, start=2):
        if not any(_cell_text(value) for value in row):
            continue
        if len(batch) >= BATCH_MAX_ROWS:
            raise HTTPException(status_code=400, detail=f"A planilha pode ter no maximo {BATCH_MAX_ROWS} carros")

        def value_for(key: str) -> Any:
            index = headers.get(key)
            return row[index] if index is not None and index < len(row) else None

        ano = _cell_int(value_for("ano"))
        ano_de = _cell_int(value_for("ano_de")) or ano
        ano_ate = _cell_int(value_for("ano_ate")) or ano
        pages = _cell_int(value_for("pages")) or 1
        per_page = _cell_int(value_for("per_page")) or 24
        try:
            request = RecommendationRequest(
                marca=_cell_text(value_for("marca")),
                modelo=_cell_text(value_for("modelo")),
                localidade=_cell_text(value_for("localidade")),
                cor=_cell_text(value_for("cor")),
                ano_de=ano_de,
                ano_ate=ano_ate,
                pages=min(max(pages, 1), 3),
                per_page=min(max(per_page, 1), 24),
            )
            _validate_year_range(request.ano_de, request.ano_ate)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Linha {excel_row_number} invalida: {exc}") from exc
        batch.append((excel_row_number, request))

    if not batch:
        raise HTTPException(status_code=400, detail="A planilha nao tem carros para pesquisar")
    return batch


def _autosize_sheet(sheet: Any) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 58)
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _workbook_bytes(workbook: Any) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _batch_template_workbook() -> BytesIO:
    Workbook, _, Font = _xlsx_tools()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "carros"
    sheet.append(["marca", "modelo", "localidade", "cor", "ano_de", "ano_ate"])
    sheet.append(["Renault", "Sandero", "RJ", "", 2019, 2022])
    sheet.append(["Honda", "City", "SP", "Branco", 2020, 2022])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    _autosize_sheet(sheet)
    return _workbook_bytes(workbook)


def _is_retryable_webmotors_error(exc: Exception) -> bool:
    response = getattr(exc, "response", None)
    status_code = getattr(response, "status_code", None)
    if status_code in (401, 403):
        return True
    message = str(exc).lower()
    return "401 client error" in message or "403 client error" in message


def _collect_batch_analysis_with_retry(
    request: RecommendationRequest,
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]], dict[str, Any] | None]:
    last_error: Exception | None = None
    for attempt in range(1, 5):
        try:
            return _collect_recommendation_analysis(request)
        except Exception as exc:
            if not _is_retryable_webmotors_error(exc):
                raise
            last_error = exc
            if attempt == 4:
                break
            get_webmotors_client().ensure_session(force=True, clear_profile=True)
            time.sleep(min(attempt * 2, 6))
    assert last_error is not None
    raise last_error


def _batch_results_workbook(batch: list[tuple[int, RecommendationRequest]]) -> BytesIO:
    Workbook, _, Font = _xlsx_tools()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "resultado"
    headers = [
        "linha_origem",
        "marca",
        "modelo",
        "localidade",
        "cor",
        "ano_de",
        "ano_ate",
        "status",
        "total_anuncios_filtrados",
        "total_encontrados_webmotors",
        "total_ignorados_por_localidade",
        "preco_medio",
        "km_medio",
        "titulo_melhor",
        "preco_melhor",
        "km_melhor",
        "ano_modelo",
        "ano_fabricacao",
        "cor_melhor",
        "cidade",
        "estado",
        "url_anuncio",
        "erro",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row_number, request in batch:
        try:
            busca, metricas, _, melhor = _collect_batch_analysis_with_retry(request)
            status = "ok" if melhor else "sem resultado"
            sheet.append([
                row_number,
                busca["marca"],
                busca["modelo"],
                busca["localidade"],
                busca["cor"],
                busca["ano_de"],
                busca["ano_ate"],
                status,
                metricas["total_anuncios"],
                metricas["total_anuncios_encontrados_webmotors"],
                metricas["total_ignorados_por_localidade"],
                metricas["media_preco"],
                metricas["media_km"],
                (melhor or {}).get("titulo"),
                (melhor or {}).get("preco"),
                (melhor or {}).get("km"),
                (melhor or {}).get("ano_modelo"),
                (melhor or {}).get("ano_fabricacao"),
                (melhor or {}).get("cor"),
                (melhor or {}).get("cidade"),
                (melhor or {}).get("estado"),
                (melhor or {}).get("url"),
                "",
            ])
            link_cell = sheet.cell(row=sheet.max_row, column=headers.index("url_anuncio") + 1)
            if link_cell.value:
                link_cell.hyperlink = link_cell.value
                link_cell.style = "Hyperlink"
        except Exception as exc:
            sheet.append([
                row_number,
                request.marca,
                request.modelo,
                request.localidade,
                request.cor,
                request.ano_de,
                request.ano_ate,
                "erro",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                str(exc),
            ])

    _autosize_sheet(sheet)
    return _workbook_bytes(workbook)


# ==========================================
# Endpoint Tavily (busca aproximada via snippets) - legado
# ==========================================

@app.get("/search_vehicle")
def search_vehicle(
    marca: str,
    modelo: str,
    ano: int,
    cor: str = "",
    localidade: str = ""
):

    query = f"""
    site:webmotors.com.br
    {marca}
    {modelo}
    {ano}
    {cor}
    {localidade}
    veículo usado
    preço
    quilometragem
    """

    result = get_tavily_client().search(
        query=query,
        search_depth="advanced",
        max_results=50
    )

    webmotors_results = keep_only_webmotors(
        result.get("results", [])
    )

    anuncios_extraidos = []

    for item in webmotors_results:

        content = item.get(
            "content",
            ""
        )

        anuncios = extract_vehicle_data(
            content
        )

        for anuncio in anuncios:

            anuncios_extraidos.append({

                "titulo": item.get(
                    "title",
                    ""
                ),

                "url": item.get(
                    "url",
                    ""
                ),

                "preco": anuncio["preco"],

                "km": anuncio["km"],

                "localidade_pesquisada": localidade,

                "cor_pesquisada": cor
            })

    medias = calcular_medias(anuncios_extraidos)

    return {
        "marca": marca,
        "modelo": modelo,
        "ano": ano,
        "cor": cor,
        "localidade": localidade,
        "total_anuncios": len(anuncios_extraidos),
        "media_preco": medias["media_preco"],
        "media_km": medias["media_km"],
        "anuncios": anuncios_extraidos,
    }


# ==========================================
# Endpoints Webmotors (dados REAIS via API interna)
#
# Passa pelo PerimeterX abrindo o Chrome real uma vez (mint), depois usa
# HTTP puro com os cookies colhidos. Ver webmotors_scraper.py.
# ==========================================

@app.get("/webmotors/search")
def webmotors_search(
    marca: str,
    modelo: str = "",
    localidade: str = Query("", description="Localidade da busca, ex.: SP, São Paulo ou São Paulo/SP"),
    cor: str = Query("", description="Cor do veículo, ex.: Branco, Preto, Prata"),
    ano_de: int | None = Query(None, description="Ano inicial do modelo, ex.: 2019"),
    ano_ate: int | None = Query(None, description="Ano final do modelo, ex.: 2022"),
    page: int = 1,
    per_page: int = 24,
):
    """Busca anúncios reais na Webmotors (preço/km/cor/cidade corretos)."""
    if ano_de is not None and ano_ate is not None and ano_de > ano_ate:
        raise HTTPException(status_code=400, detail="ano_de não pode ser maior que ano_ate")

    extra = wm._build_search_extra(
        localidade=localidade,
        cor=cor,
        ano_de=ano_de,
        ano_ate=ano_ate,
    )

    try:
        data = get_webmotors_client().search(
            make=marca, model=modelo, page=page, per_page=per_page, extra=extra
        )
    except wm.WebmotorsBlocked as exc:
        raise HTTPException(status_code=502, detail=str(exc))

    anuncios = data.get("SearchResults", [])
    anuncios_normalizados = [wm.normalize_detail(it) for it in anuncios]
    medias = calcular_medias(anuncios_normalizados)

    return {
        "marca": marca,
        "modelo": modelo,
        "localidade": localidade,
        "cor": cor,
        "ano_de": ano_de,
        "ano_ate": ano_ate,
        "page": page,
        "total_disponivel": data.get("Count"),
        "total_anuncios": len(anuncios),
        "media_preco": medias["media_preco"],
        "media_km": medias["media_km"],
        "anuncios": anuncios,
        "payload_completo": data,
    }


@app.get("/webmotors/detail")
def webmotors_detail(
    marca: str,
    modelo: str = "",
    localidade: str = Query("", description="Localidade da busca, ex.: SP, São Paulo ou São Paulo/SP"),
    cor: str = Query("", description="Cor do veículo, ex.: Branco, Preto, Prata"),
    ano_de: int | None = Query(None, description="Ano inicial do modelo, ex.: 2019"),
    ano_ate: int | None = Query(None, description="Ano final do modelo, ex.: 2022"),
    pages: int = 1,
    per_page: int = 12,
):
    """Busca + detalha cada usado (payload completo do /api/detail)."""
    if ano_de is not None and ano_ate is not None and ano_de > ano_ate:
        raise HTTPException(status_code=400, detail="ano_de não pode ser maior que ano_ate")

    extra = wm._build_search_extra(
        localidade=localidade,
        cor=cor,
        ano_de=ano_de,
        ano_ate=ano_ate,
    )

    try:
        details = get_webmotors_client().iter_details(
            make=marca, model=modelo, pages=pages, per_page=per_page, extra=extra
        )
    except wm.WebmotorsBlocked as exc:
        raise HTTPException(status_code=502, detail=str(exc))

    anuncios_normalizados = [wm.normalize_detail(d) for d in details]
    medias = calcular_medias(anuncios_normalizados)

    return {
        "marca": marca,
        "modelo": modelo,
        "localidade": localidade,
        "cor": cor,
        "ano_de": ano_de,
        "ano_ate": ano_ate,
        "total_anuncios": len(details),
        "media_preco": medias["media_preco"],
        "media_km": medias["media_km"],
        "anuncios": details,
    }


@app.post("/recommendation", response_model=RecommendationResponse)
def purchase_recommendation(request: RecommendationRequest):
    """Busca anúncios reais, envia o resumo ao Bedrock Agent e retorna a recomendação."""
    try:
        busca, metricas, anuncios, melhor_opcao = _collect_recommendation_analysis(request)
    except wm.WebmotorsBlocked as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    prompt = _build_agent_prompt(
        busca=busca,
        metricas=metricas,
        anuncios=anuncios,
        melhor_opcao_local=melhor_opcao,
    )

    try:
        recomendacao = _invoke_bedrock_agent(prompt)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Erro ao chamar Bedrock Agent: {exc}") from exc

    return {
        "busca": busca,
        "metricas": metricas,
        "melhor_opcao_local": melhor_opcao,
        "recomendacao": recomendacao,
        "anuncios_analisados": anuncios,
    }


@app.get("/recommendation/batch/template")
def batch_template():
    """Baixa um modelo XLSX para pesquisa em lote."""
    return StreamingResponse(
        _batch_template_workbook(),
        media_type=BATCH_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="modelo_pesquisa_carros.xlsx"'},
    )


@app.post("/recommendation/batch")
def batch_purchase_recommendation(file: UploadFile = File(...)):
    """Recebe ate 20 carros em XLSX e devolve outro XLSX com o melhor anuncio de cada busca."""
    content = file.file.read()
    batch = _read_batch_requests(content)
    workbook = _batch_results_workbook(batch)
    return StreamingResponse(
        workbook,
        media_type=BATCH_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": 'attachment; filename="resultado_recomendacoes.xlsx"',
            "X-Batch-Items": str(len(batch)),
        },
    )


@app.get("/webmotors/detail_by_url")
def webmotors_detail_by_url(url: str):
    """Detalhe de um anúncio a partir da URL de detail (/api/detail/...)."""
    try:
        detail = get_webmotors_client().get_detail(url=url)
    except wm.WebmotorsBlocked as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    return detail


@app.post("/webmotors/refresh_session")
def webmotors_refresh_session(clear_profile: bool = False):
    """Força um novo mint (abre o Chrome e resolve o PerimeterX de novo)."""
    try:
        sess = get_webmotors_client().ensure_session(force=True, clear_profile=clear_profile)
    except wm.WebmotorsBlocked as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    return {"ok": True, "cookies": list(sess.cookies), "user_agent": sess.user_agent}
